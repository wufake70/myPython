# _*_coding :utf-8 _*_
# @Time :2022/7/25 23:18
# @File : python第三方库_openpyxl
# @Project : python_NCRE

from pprint import pprint

import openpyxl

# 读取 excel文档 即 工作簿
excel = openpyxl.load_workbook(r"C:\Users\yui\AppData\Roaming\Tencent\WeChatAppStore\WeChatAppStore "
                               r"Files\wxid_4eyihzpaijud22\Files\001.xlsx")

# 获取 工作簿中的 全部工作表的名字
all_sheet = excel.sheetnames  # sheet_ = excel.get_sheet_names 已弃用
print(all_sheet)  # ['Sheet1', 'Sheet2', 'Sheet3']

# 获取 sheet 的名字
# sheet_1 = all_sheet[0]        列表
# print(sheet_1.title())        字符串的title（） 方法 首字母 大写
# sheet_1 = excel.get_sheet_by_name('Sheet1')    已弃用
sheet_1 = excel['Sheet1']  # 工作表对象
print(sheet_1.title)

# 获取 当前活动表
sheet_act = excel.active
print(sheet_act.title)

# #从表中 获取单元格 cell
print(sheet_1['A1'])  # 获取 单元格 column A、row 1 的对象
# print(sheet_1['A'])         # 获取 column 为A 的一列数据 并保存为 tuple
print(sheet_1['A'][:100:2])  # 获取 …… 通过索引 取出
# 获取 相关值 并保存
cell_A_values = [i.value for i in sheet_1['A'][:100:10]]
pprint(cell_A_values)

print(sheet_1['1'])  # 获取 表的 第一行的 所有 cell
cell_1_values = [i.value for i in sheet_1['1']]
print(cell_1_values)

print("*"*50, end='\n\n\n')
# 华而不实
# cell_columns = [i for i in range(1, len(sheet_1['A']+1))]
for i in range(1, len(sheet_1["A"])+1):
    print([i.value for i in sheet_1[f'{i}']])

# 指定 数值 row 行；column 列
cell = sheet_1.cell(row=1, column=1).value
# print(cell)

print("*"*50, end='\n\n\n')
# 获取 excel 矩形区域的数据
cells = sheet_1['A1':'G10']
# print(cells)        # 每一行 cells 为一组 元祖 ，所有行的 元祖 又用 元祖存储

for i in sheet_1:
    for column in i:
        print(column.value, end='\t')
    print('')

print("*"*50, end='\n\n\n')
# 创建 并保存 excel文档
new_excel = openpyxl.Workbook()
sheet_act = new_excel.active
sheet_act.title = '新表格'

# 创建/ 删除 表，index 为表格 添加索引，title 表格 名字
new_excel.create_sheet(index=1, title='新表2')
new_excel.create_sheet(index=2, title='新表3')

# 填充 数据 同上 读取数据 ，注意
sheet_write = new_excel['新表3']
for i in range(1, len(sheet_1["A"])+1):
    for column in 'ABCDEFG':
        sheet_write[rf'{column}{i}'].value = sheet_1[f'{column}{i}'].value


new_excel.save('新的excel.xlsx')                         























