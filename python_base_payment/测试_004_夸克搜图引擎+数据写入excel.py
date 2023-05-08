import pprint
import re
import json
import requests
import openpyxl


def base_url():
    query = input('请键入关键字:')
    start = input('请键入获取数值:')
    url = f'https://vt.sm.cn/api/pic/list?query={query}&tag=&limit=20&start={start}&databucket=default&ad=on&hid' \
          f'=ZWwq0EVuH0F4ykI6Y5dSluXLlvw7hhdz&from=wm850032&fr=&uid=7224ce7c4ca4fa9734fd3a904611d13f%7C54691003951%7C' \
          f'%7C1647355696 '
    return url, start, query


def request_1(url):
    # 该url只返回20个图片数据 url = f'https://vt.sm.cn/api/pic/list?query={
    # query}&tag=&limit=20&start=0&databucket=default&ad=on&hid=ZWwq0EVuH0F4ykI6Y5dSluXLlvw7hhdz&from=wm850032&fr
    # =&uid=7224ce7c4ca4fa9734fd3a904611d13f%7C54691003951%7C%7C1647355696'

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:102.0) Gecko/20100101 Firefox/102.0",
        "Accept-Language": "zh-CN,zh;q=0.8,zh-TW;q=0.7,zh-HK;q=0.5,en-US;q=0.3,en;q=0.2",
        "DNT": "1",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Origin": "https://vt.quark.cn",
        "Referer": "https://vt.quark.cn/"
    }
    response_1 = requests.get(url, headers)

    return response_1


dic_1 = []


def img_dic(response_1, dic_1=[]):                                # 使用json模块加载数据
    # print(response.text)                              # 字符串类型
    response_1 = json.loads(response_1.text)              # 将字符串对象加载成为 python 对象
    # pprint.pprint(response_1)                         # 查看字典数据结构 方便获取资源url
    dic = response_1['data']['hit']['imgInfo']['item']  # itme 对应的值为列表类型
    # pprint.pprint(dic)
    dic_1 += dic
    return dic_1


def save_ex(dic):
    # print(dic)
    wb = openpyxl.Workbook()                            # 创建一个空的工作簿对象
    sheet = wb.active
    sheet.title = '镁铝'
    sheet['A1'].value = '主题'
    sheet['A2'].value = '序号'
    sheet['B2'].value = '标题'
    sheet['C2'].value = '链接'
    row = 3                                             # 初始行

    for i in dic:
        column = 1                                      # 初始列
        sheet.cell(row, column).value = '%04d' % (dic.index(i) + 1)   # 添加序号
        sheet.cell(row, column + 1).value = i['title']          # 写入名字
        sheet.cell(row, column + 2).value = i['bigPicCache']    # 写入链接
        # pprint.pprint(i['title'])                     # 图片的名字
        # pprint.pprint(i['bigPicCache'])               # 图片的url
        row += 1                                        # 增加行数
    wb.save(f'美女{i["title"]}.xlsx')


def moro_url(url, start, query):
    for i in range(int(start) // 20):
        start = 20 * i + 1
        url = f'https://vt.sm.cn/api/pic/list?query={query}&tag=&limit=20&start={start}&databucket=default&ad=on&hid' \
          f'=ZWwq0EVuH0F4ykI6Y5dSluXLlvw7hhdz&from=wm850032&fr=&uid=7224ce7c4ca4fa9734fd3a904611d13f%7C54691003951%7C' \
          f'%7C1647355696 '
        response = request_1(url)
        dic = img_dic(response)
    # print(len(dic))
    # eval(input(':'))
    save_ex(dic)


a = base_url()
moro_url(a[0], a[1], a[2])






















































